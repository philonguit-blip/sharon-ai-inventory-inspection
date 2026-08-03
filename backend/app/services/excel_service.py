"""Create a KiotViet purchase-import workbook from bakery detections."""

from __future__ import annotations

import os
from collections import OrderedDict
from copy import copy
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from app.config import TEMPLATE_PATH


REQUIRED_HEADERS = {
    "product_code": "Mã hàng",
    "product_name": "Tên hàng",
    "purchase_price": "Đơn giá",
    "quantity": "Số lượng",
}


class ExcelExportError(RuntimeError):
    """Raised when a safe KiotViet import workbook cannot be generated."""


class ExcelService:
    """Aggregate inference results and fill the KiotViet Excel template."""

    def __init__(self, template_path: Path | str = TEMPLATE_PATH) -> None:
        self.template_path = Path(template_path).expanduser().resolve()
        if not self.template_path.is_file():
            raise ExcelExportError(f"Excel template not found: {self.template_path}")

    def aggregate_products(
        self, inference_results: Iterable[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        """Sum quantities by KiotViet product code across multiple images."""
        grouped: OrderedDict[str, dict[str, Any]] = OrderedDict()
        result_count = 0

        for result in inference_results:
            result_count += 1
            image_name = str(result.get("image_name") or f"image_{result_count}")
            status = str(result.get("status", "SUCCESS")).upper()
            if status != "SUCCESS":
                message = str(result.get("error") or "unknown inference error")
                raise ExcelExportError(
                    f"Cannot export because inference failed for {image_name}: {message}"
                )

            products = result.get("products")
            if not isinstance(products, list):
                raise ExcelExportError(
                    f"Inference result for {image_name} has no valid products list."
                )

            for product in products:
                self._add_product(grouped, product, image_name)

        if result_count == 0:
            raise ExcelExportError("At least one inference result is required.")
        if not grouped:
            raise ExcelExportError(
                "No products were detected; an empty KiotViet import file was not created."
            )
        return list(grouped.values())

    @staticmethod
    def _add_product(
        grouped: OrderedDict[str, dict[str, Any]],
        product: Any,
        image_name: str,
    ) -> None:
        if not isinstance(product, dict):
            raise ExcelExportError(
                f"Invalid product data in inference result for {image_name}."
            )

        product_code = str(product.get("product_code") or "").strip()
        product_name = str(product.get("product_name") or "").strip()
        if not product_code:
            raise ExcelExportError(f"Missing product code for {image_name}.")
        if not product_name:
            raise ExcelExportError(
                f"Missing product name for product {product_code} in {image_name}."
            )

        raw_quantity = product.get("quantity")
        if isinstance(raw_quantity, bool):
            raise ExcelExportError(
                f"Invalid quantity for product {product_code} in {image_name}."
            )
        try:
            quantity_number = float(raw_quantity)
        except (TypeError, ValueError) as exc:
            raise ExcelExportError(
                f"Invalid quantity for product {product_code} in {image_name}."
            ) from exc
        if quantity_number <= 0 or not quantity_number.is_integer():
            raise ExcelExportError(
                f"Quantity must be a positive integer for product {product_code}."
            )
        quantity = int(quantity_number)

        existing = grouped.get(product_code)
        if existing is None:
            grouped[product_code] = {
                "product_code": product_code,
                "product_name": product_name,
                # The agreed KiotViet purchase price for this pipeline is 0.
                "purchase_price": 0,
                "quantity": quantity,
            }
            return

        if existing["product_name"] != product_name:
            raise ExcelExportError(
                "Conflicting product names for code "
                f"{product_code}: {existing['product_name']!r} and {product_name!r}."
            )
        existing["quantity"] += quantity

    def create_import_workbook(
        self,
        inference_results: Iterable[dict[str, Any]],
        output_path: Path | str,
    ) -> dict[str, Any]:
        """Generate and validate one KiotViet purchase-import workbook."""
        products = self.aggregate_products(inference_results)
        destination = Path(output_path).expanduser().resolve()
        if destination.suffix.lower() != ".xlsx":
            raise ExcelExportError("The output path must end with .xlsx.")
        if destination == self.template_path:
            raise ExcelExportError("The source Excel template cannot be overwritten.")

        try:
            workbook = load_workbook(self.template_path)
            worksheet = workbook["PurchaseOrderTemplate"]
        except Exception as exc:
            raise ExcelExportError(
                f"Cannot open Excel template: {self.template_path}"
            ) from exc

        header_columns = self._read_header_columns(worksheet)
        template_row = self._find_template_row(worksheet)
        original_last_row = max(worksheet.max_row, 2)
        output_last_row = len(products) + 1
        final_last_row = max(original_last_row, output_last_row)

        # Remove example values while preserving the template's cell styles.
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=final_last_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.value = None

        for row_number, product in enumerate(products, start=2):
            if row_number > original_last_row:
                self._copy_row_style(
                    worksheet,
                    source_row=template_row,
                    target_row=row_number,
                )

            code_cell = worksheet.cell(
                row=row_number, column=header_columns["product_code"]
            )
            code_cell.value = product["product_code"]
            code_cell.number_format = "@"

            worksheet.cell(
                row=row_number, column=header_columns["product_name"]
            ).value = product["product_name"]

            price_cell = worksheet.cell(
                row=row_number, column=header_columns["purchase_price"]
            )
            price_cell.value = 0
            price_cell.number_format = "#,##0"

            quantity_cell = worksheet.cell(
                row=row_number, column=header_columns["quantity"]
            )
            quantity_cell.value = product["quantity"]
            quantity_cell.number_format = "#,##0"

        # Delete unused sample rows so no stale blank rows are part of the table.
        if original_last_row > output_last_row:
            worksheet.delete_rows(
                output_last_row + 1,
                original_last_row - output_last_row,
            )

        self._widen_required_columns(worksheet, header_columns, products)
        self._resize_table(worksheet, output_last_row)
        destination.parent.mkdir(parents=True, exist_ok=True)
        temporary = destination.with_name(
            f".{destination.stem}.{uuid4().hex}.tmp.xlsx"
        )

        try:
            workbook.save(temporary)
            os.replace(temporary, destination)
        except Exception as exc:
            if temporary.exists():
                temporary.unlink()
            raise ExcelExportError(f"Cannot save Excel file: {destination}") from exc
        finally:
            workbook.close()

        self._validate_saved_workbook(destination, products)
        return {
            "path": str(destination),
            "product_count": len(products),
            "total_quantity": sum(int(item["quantity"]) for item in products),
            "products": products,
        }

    @staticmethod
    def _read_header_columns(worksheet: Any) -> dict[str, int]:
        by_title: dict[str, int] = {}
        for cell in worksheet[1]:
            if cell.value is not None:
                by_title[str(cell.value).strip()] = int(cell.column)

        missing = [
            title for title in REQUIRED_HEADERS.values() if title not in by_title
        ]
        if missing:
            raise ExcelExportError(
                "The Excel template is missing required columns: "
                + ", ".join(missing)
            )
        return {
            key: by_title[title] for key, title in REQUIRED_HEADERS.items()
        }

    @staticmethod
    def _find_template_row(worksheet: Any) -> int:
        for row_number in range(2, worksheet.max_row + 1):
            if any(
                worksheet.cell(row=row_number, column=column).has_style
                for column in range(1, worksheet.max_column + 1)
            ):
                return row_number
        return 2

    @staticmethod
    def _copy_row_style(
        worksheet: Any, source_row: int, target_row: int
    ) -> None:
        for column in range(1, worksheet.max_column + 1):
            source: Cell = worksheet.cell(row=source_row, column=column)
            target: Cell = worksheet.cell(row=target_row, column=column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[
            source_row
        ].height

    @staticmethod
    def _widen_required_columns(
        worksheet: Any,
        header_columns: dict[str, int],
        products: list[dict[str, Any]],
    ) -> None:
        content_by_key = {
            "product_code": [str(item["product_code"]) for item in products],
            "product_name": [str(item["product_name"]) for item in products],
        }
        maximum_widths = {"product_code": 28.0, "product_name": 64.0}

        for key, values in content_by_key.items():
            column_index = header_columns[key]
            letter = get_column_letter(column_index)
            current_width = float(worksheet.column_dimensions[letter].width or 0)
            desired_width = max(
                len(REQUIRED_HEADERS[key]) + 2,
                max(len(value) for value in values) + 2,
            )
            worksheet.column_dimensions[letter].width = min(
                maximum_widths[key], max(current_width, float(desired_width))
            )

    @staticmethod
    def _resize_table(worksheet: Any, last_row: int) -> None:
        tables: list[Table] = list(worksheet.tables.values())
        if not tables:
            raise ExcelExportError("The Excel template contains no import table.")
        if len(tables) != 1:
            raise ExcelExportError(
                "The Excel template must contain exactly one import table."
            )
        tables[0].ref = f"A1:M{last_row}"

    @classmethod
    def _validate_saved_workbook(
        cls, path: Path, expected_products: list[dict[str, Any]]
    ) -> None:
        try:
            workbook = load_workbook(path, read_only=False, data_only=True)
            worksheet = workbook["PurchaseOrderTemplate"]
            header_columns = cls._read_header_columns(worksheet)
            actual: list[dict[str, Any]] = []
            for row_number in range(2, len(expected_products) + 2):
                actual.append(
                    {
                        "product_code": str(
                            worksheet.cell(
                                row=row_number,
                                column=header_columns["product_code"],
                            ).value
                            or ""
                        ),
                        "product_name": str(
                            worksheet.cell(
                                row=row_number,
                                column=header_columns["product_name"],
                            ).value
                            or ""
                        ),
                        "purchase_price": worksheet.cell(
                            row=row_number,
                            column=header_columns["purchase_price"],
                        ).value,
                        "quantity": worksheet.cell(
                            row=row_number,
                            column=header_columns["quantity"],
                        ).value,
                    }
                )
        except Exception as exc:
            raise ExcelExportError(f"Cannot validate generated Excel file: {path}") from exc
        finally:
            if "workbook" in locals():
                workbook.close()

        if actual != expected_products:
            raise ExcelExportError(
                "Generated Excel content does not match the aggregated detections."
            )
